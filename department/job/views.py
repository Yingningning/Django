from django.shortcuts import render,redirect,HttpResponse
from django.http import JsonResponse
from job import models
import json
from job.utils.Pagenation import Pagenation
# Create your views here.
def depart_list(request):
    
    titles = models.Department.objects.all()
    return render(request,'depart_list.html',{'titles':titles})
    
    
def depart_add(request):
    if request.method == "GET":
        return render(request,'depart_add.html')
    else:
        title = request.POST.get('title')
        models.Department.objects.create(title = title)
        return redirect('/depart_list')
    
def depart_delete(request):
    
    uid = request.GET.get('uid')
    models.Department.objects.filter(id = uid).delete()
    return redirect('/depart_list')

def depart_edit(request,nid):
    if request.method == "GET":
        depart = models.Department.objects.filter(id = nid).first()
        return render(request,'depart_edit.html',{'depart':depart})
    else:
        new_title = request.POST.get('title')
        models.Department.objects.filter(id = nid).update(title = new_title)
        return redirect('/depart_list')
    
from openpyxl import load_workbook
def depart_mutil(request):
    file_obj = request.FILES.get('file')
    print(type(file_obj))
    
    #将文件对象传递给openpyxl,读取文件内容
    wb = load_workbook(file_obj)
    sheet = wb.worksheets[0]
    
    #循环获取每一行数据，并保存到数据库
    for row in sheet.iter_rows(min_row=2):
        title = row[0].value
        exists = models.Department.objects.filter(title = title).exists()
        if not exists:
            print(row[0].value)
            models.Department.objects.create(title = title)
    
    return redirect('/depart_list')
    
def user_list(request):
    user_info = models.User_info.objects.all()
    
    # user.depart.title 根据外键获取部门名称
    # user.get_gender_display() # 根据性别字段获取中文名
    return render(request,'user_list.html',{'user_info':user_info})


from django import forms
class UserModelForm(forms.ModelForm):
    
    name = forms.CharField(min_length=2,label="姓名")
    # name = forms.CharField(min_length=2,label="姓名",widget=forms.TextInput(attrs={"class":"form-control"}))
    class Meta:
        model = models.User_info
        
        fields = ["name","pwd","age","account","create_time","gender","depart"]
        # widgets = {
        #     "name" : forms.TextInput(attrs={"class":"form-control"})
        # }
    def __init__(self,*args,**kwargs):
        super().__init__(*args,**kwargs)
        
        for name,field in self.fields.items():
            field.widget.attrs = {"class":"form-control","placeholder":field.label}
            
def user_add(request):
    """基于modelForm的添加用户"""
    if request.method == "GET":
    
        form = UserModelForm()
        return render(request,"change.html",{'form':form,'title':"添加用户"})
    
    form = UserModelForm(data=request.POST)
    if form.is_valid():
        form.save()
        return redirect('/user_list')
    return render(request,"change.html",{'form':form,'title':"添加用户"})

def user_edit(request,nid):
    user = models.User_info.objects.filter(id = nid).first()
    if request.method == "GET":
        # 根据id获取用户信息
        form = UserModelForm(instance = user)
        return render(request,'user_edit.html',{'form':form})
    
    form =UserModelForm(data=request.POST,instance=user)
    if form.is_valid():
        form.save()
        return redirect('/user_list')
    return render(request,'user_edit.html',{'form':form})

def user_delete(request,nid):
    models.User_info.objects.filter(id = nid).delete()
    return redirect('/user_list')


def Admin_list(request):
    # models.Admin_info.objects.create(username = 'ynn',pwd = '12345')
    # info = request.session.get('info')
    # if not info:
    #     return redirect('/login')
    
    
    
    data_dict = {}
    search_key = request.GET.get('q','')
    if search_key:
        data_dict['username__contains'] = search_key
        queryset = models.Admin_info.objects.filter(**data_dict)
    else:
        queryset = models.Admin_info.objects.all()
    page_str = Pagenation(request,queryset).html()
    context = {
        'queryset':queryset,
        'page_str':page_str
    }
    return render(request,'Admin_list.html',context)

from job.utils.modelforms import AdminModelFrom,AdminEditModelFrom,Account,TaskForm
def admin_add(request): 
    title = "添加管理员"
    if request.method =="GET":
        form = AdminModelFrom()
        return render(request,'change.html',{'form':form,'title':title})
    form = AdminModelFrom(data = request.POST)
    if form.is_valid():
        form.save()
        return redirect('/Admin_list')
    return render(request,'change.html',{'form':form,'title':title})
    
    
def admin_edit(request,nid):
    title = "编辑管理员"
    admin = models.Admin_info.objects.filter(id = nid).first()
    if not admin:
        return redirect('/Admin_list')
    if request.method == "GET":
        form = AdminEditModelFrom(instance = admin)
        return render(request,'change.html',{'form':form,'title':title})
    form = AdminEditModelFrom(data = request.POST,instance = admin)
    if form.is_valid():
        form.save()
        return redirect('/Admin_list')
    
 

def login(request):
    if request.method =="GET":
        form = Account()
        return render(request,'login.html',{"form":form})
    form = Account(data = request.POST)
    if form.is_valid():
        # print(form.cleaned_data)
        user_input_code = form.cleaned_data.pop('code')
        true_code = request.session.get('img_code','')
        if user_input_code.lower() != true_code.lower():
            form.add_error('code','验证码错误')
            return render(request,'login.html',{"form":form})
        
        obj = models.Admin_info.objects.filter(**form.cleaned_data).first()
        if not obj:
            form.add_error("pwd","用户名或密码错误")
            return render(request,'login.html',{"form":form})
        
        # 正确后，生成随机字符串，写到浏览器的cookie中，并存入session，跳转到首页
        # request.session['info'] = obj.username
        request.session['info'] = {'id':obj.id,'username':obj.username}
        request.session.set_expiry(60*60*24*7)
        return redirect('/Admin_list')
    return render(request,'login.html',{"form":form})

def logout(request):
    request.session.clear()
    return redirect('/login')


from job.utils.code import create_code
from io import BytesIO
def img_code(request):
    img,code_str = create_code()
    # print(code_str)
    request.session['img_code'] = code_str
    # 设置session的有效期为60秒
    request.session.set_expiry(60)
    
    stream = BytesIO()
    img.save(stream,'png')
    return HttpResponse(stream.getvSalue())

def task_list(request):
    form = TaskForm()
    queryset = models.Task.objects.all()
    return render(request,'task_list.html',{'form':form,'queryset':queryset})
    
from django.views.decorators.csrf import csrf_exempt


@csrf_exempt # 取消csrf校验
def task_ajax(request):
    print(request.POST)
    
    data_dicr = {'status':'true'}
    return JsonResponse(data_dicr)

@csrf_exempt
def task_add(request):
    # print(request.POST)
    # data_dicr = {'status':'true'}
    # return JsonResponse(data_dicr)
    form = TaskForm(data=request.POST)
    if form.is_valid():
        form.save()
        data_dict = {'status':True}

        return JsonResponse(data_dict)
    data_dict = {'status':False,'msg':form.errors}
    return JsonResponse(data_dict)

from job.utils.modelforms import OrderForm
def order_list(request):
    querset = models.Order.objects.all()
    page_str = Pagenation(request,querset).html()
    form = OrderForm()
    entext = {
        'form':form,
        'queryset':querset,
        'page_str':page_str
    }
    return render(request,'order_list.html',entext)

@csrf_exempt
def order_add(request):
    form = OrderForm(data= request.POST)
    if form.is_valid():
        form.instance.oid = 'xxxxx'
        form.instance.admin_id = request.session['info']['id']
        
        form.save()
        return JsonResponse({'status':True})
    return JsonResponse({'status':False,'msg':form.errors})
    
def order_del(request):
    id = request.GET.get('uid')
    models.Order.objects.filter(id = id).delete()
    status_dict = {'status':True}
    return JsonResponse(status_dict)

def order_detail(request):
    uid = request.GET.get('uid')
    # 这种形式返回的是一个字典
    row_obj = models.Order.objects.filter(id = uid).values('title','price','status').first()
    if not row_obj:
        return JsonResponse({'status':False,'msg':'订单不存在'})
    
    result = {
        'status': True,
        'data': row_obj
    }
    return JsonResponse(result)

@csrf_exempt 
def order_edit(request):
    uid = request.GET.get('uid')
    obj = models.Order.objects.filter(id = uid).first()
    if not obj:
        return JsonResponse({'status':False,'tips':'订单不存在'})
    form = OrderForm(data=request.POST,instance=obj)
    if form.is_valid():
        form.save()
        return JsonResponse({'status':True})
    return JsonResponse({'status':False,'msg':form.errors})


def chart_list(request):
    return render(request,'chart_list.html')

def chart_bar(request):
    # 构造数据 数据可以从数据库中获取
    legent = ['销量', '业绩']
    data_list = [
         {
                    "name": '销量',
                    "type": 'bar',
                    "data": [5, 20, 36, 10, 10, 20]
                },
                {
                    "name": '业绩',
                    "type": 'bar',
                    "data": [23, 13, 15, 32, 17, 14]
                }
    ]
    data =  ['衬衫', '羊毛衫', '雪纺衫', '裤子', '高跟鞋', '袜子']
    
    result = {
        'status': True,
        'data':{
            "legent":legent,
            "data_list":data_list,
            "data" :data
        }
    }
    
    return JsonResponse(result)

def upload_list(request):
    if request.method == "GET":
        return render(request,'upload_list.html')
    # request.POST 请求体中的数据
    # request.FILES 上传的文件对象
    file = request.FILES.get('file') #文件对象
    print(file.name) #文件名
    
    f = open('D:/APP/VS CODE/web/department/job/static/upload/'+file.name,'wb')
    for chunk in file.chunks():
        f.write(chunk)
    f.close()
    return HttpResponse("上传成功")

from job.utils.modelforms import UpModelForm
def upload_modelform(request):
    # 基于modelForm的上传文件
    if request.method == "GET":
        form = UpModelForm()
        return render(request,'upload_modelform.html',{'form':form,'title':'上传文件'})
    form = UpModelForm(data = request.POST,files = request.FILES)
    if form.is_valid():
        # 对于文件，自动上传，且路径写入数据库
        form.save()
        return HttpResponse("上传成功")
    return render(request,'upload_modelform.html',{'form':form,'title':'上传文件'})

def city_list(request):
    
    queryset = models.City.objects.all()
    return render(request,'city_list.html',{'queryset':queryset})

def city_add(request):
    title = "添加城市"
    if request.method == "GET":
        form = UpModelForm()
        return render(request,'upload_modelform.html',{'form':form,'title':title})
    form = UpModelForm(data = request.POST,files = request.FILES)
    if form.is_valid():
        form.save()
        return redirect('/city/list')
    
    #如果有错误信息，返回错误信息
    return render(request,'upload_modelform.html',{'form':form,'title':title})
        
    